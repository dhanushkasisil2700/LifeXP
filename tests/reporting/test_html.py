"""Tests for HTMLReport and ExcelReport (lifexp.reporting).

Checkpoint S18: Automated tests verify file creation, structure, and data
fidelity to 4 decimal places.  Visual / MathJax rendering must be inspected
manually by opening the generated files in a browser / Excel.
"""

from __future__ import annotations

import os
import re

import numpy as np
import openpyxl
import pandas as pd
import pytest

from lifexp.reporting.excel_report import ExcelReport
from lifexp.reporting.html_report import HTMLReport


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ae_df(n: int = 5) -> pd.DataFrame:
    rng = np.random.default_rng(0)
    ages = list(range(40, 40 + n))
    return pd.DataFrame({
        "age":                ages,
        "deaths":             rng.integers(1, 20, size=n).tolist(),
        "etr":                rng.uniform(50, 200, size=n).tolist(),
        "expected_ri_claims": rng.uniform(1000, 5000, size=n).tolist(),
        "actual_ri_claims":   rng.uniform(800, 6000, size=n).tolist(),
        "ae_ratio":           rng.uniform(0.6, 1.5, size=n).tolist(),
    })


def _grad_df(n: int = 5) -> pd.DataFrame:
    ages = list(range(40, 40 + n))
    return pd.DataFrame({
        "age":  ages,
        "qx":   [0.001 * (1 + 0.1 * i) for i in range(n)],
    })


# ---------------------------------------------------------------------------
# Test 1 — HTML renders and contains expected structural strings
# ---------------------------------------------------------------------------

class TestHTMLRenders:
    """HTML file is created with correct structure; MathJax tag is present."""

    @pytest.fixture(scope="class")
    def html_path(self, tmp_path_factory):
        out = str(tmp_path_factory.mktemp("html") / "study.html")
        rpt = HTMLReport(
            "Test Mortality Study",
            {"period": "2020-2022", "table_used": "IA 2012", "methodology": "CLT"},
        )
        rpt.add_ae_table(_ae_df(), caption="A/E by Age")
        rpt.add_summary_stats({"Overall A/E": "1.05", "Deaths": "234"})
        rpt.render(out)
        return out

    @pytest.fixture(scope="class")
    def html_content(self, html_path):
        with open(html_path, encoding="utf-8") as fh:
            return fh.read()

    def test_file_created(self, html_path):
        assert os.path.isfile(html_path)

    def test_doctype_present(self, html_content):
        assert "<!DOCTYPE html>" in html_content

    def test_title_in_head(self, html_content):
        assert "<title>Test Mortality Study</title>" in html_content

    def test_h1_contains_title(self, html_content):
        assert "<h1>Test Mortality Study</h1>" in html_content

    def test_mathjax_script_tag_present(self, html_content):
        """MathJax CDN script tag is embedded (Checkpoint S18 pre-condition)."""
        assert "MathJax-script" in html_content
        assert "mathjax@3" in html_content

    def test_mathjax_config_present(self, html_content):
        assert "window.MathJax" in html_content

    def test_metadata_period_present(self, html_content):
        assert "2020-2022" in html_content

    def test_metadata_table_present(self, html_content):
        assert "IA 2012" in html_content

    def test_ae_caption_present(self, html_content):
        assert "A/E by Age" in html_content

    def test_ae_formula_present(self, html_content):
        assert r"\text{A/E}" in html_content

    def test_summary_stats_values_present(self, html_content):
        assert "1.05" in html_content
        assert "234" in html_content

    def test_css_style_tag(self, html_content):
        assert "<style>" in html_content

    def test_table_class_data(self, html_content):
        assert 'class="data"' in html_content


# ---------------------------------------------------------------------------
# Test 2 — Excel file has correct sheets
# ---------------------------------------------------------------------------

class TestExcelSheets:
    """Excel file is created; all added sheets are present by name."""

    @pytest.fixture(scope="class")
    def excel_path(self, tmp_path_factory):
        out = str(tmp_path_factory.mktemp("xlsx") / "study.xlsx")
        rpt = ExcelReport()
        rpt.add_ae_sheet(_ae_df(), sheet_name="AE by Age")
        rpt.add_ae_sheet(_ae_df(3), sheet_name="AE by Gender")
        rpt.add_summary_sheet({"Overall A/E": 1.05, "Deaths": 234}, sheet_name="Summary")
        rpt.render(out)
        return out

    @pytest.fixture(scope="class")
    def wb(self, excel_path):
        return openpyxl.load_workbook(excel_path)

    def test_file_created(self, excel_path):
        assert os.path.isfile(excel_path)

    def test_ae_sheet_present(self, wb):
        assert "AE by Age" in wb.sheetnames

    def test_second_ae_sheet_present(self, wb):
        assert "AE by Gender" in wb.sheetnames

    def test_summary_sheet_present(self, wb):
        assert "Summary" in wb.sheetnames

    def test_correct_number_of_sheets(self, wb):
        assert len(wb.sheetnames) == 3

    def test_header_row_bold(self, wb):
        ws = wb["AE by Age"]
        assert ws["A1"].font.bold is True

    def test_header_text_matches_column_name(self, wb):
        ws = wb["AE by Age"]
        assert ws["A1"].value == "age"

    def test_summary_metric_column_header(self, wb):
        ws = wb["Summary"]
        assert ws["A1"].value == "Metric"

    def test_summary_value_column_header(self, wb):
        ws = wb["Summary"]
        assert ws["B1"].value == "Value"


# ---------------------------------------------------------------------------
# Test 3 — A/E table values preserved in HTML to 4 d.p.
# ---------------------------------------------------------------------------

class TestHTMLValueFidelity:
    """Table values in rendered HTML match source DataFrame to 4 decimal places."""

    DF = _ae_df(3)

    @pytest.fixture(scope="class")
    def html_content(self, tmp_path_factory):
        out = str(tmp_path_factory.mktemp("html") / "fidelity.html")
        rpt = HTMLReport("Fidelity Test", {})
        rpt.add_ae_table(self.DF, caption="Fidelity")
        rpt.render(out)
        with open(out, encoding="utf-8") as fh:
            return fh.read()

    def test_ae_ratio_values_in_html(self, html_content):
        """Each ae_ratio value appears formatted to 4 d.p. in the HTML."""
        for val in self.DF["ae_ratio"]:
            formatted = f"{val:.4f}"
            assert formatted in html_content, (
                f"Expected '{formatted}' in HTML for ae_ratio={val}"
            )

    def test_etr_values_in_html(self, html_content):
        for val in self.DF["etr"]:
            formatted = f"{val:.4f}"
            assert formatted in html_content

    def test_integer_deaths_in_html(self, html_content):
        for val in self.DF["deaths"]:
            assert str(val) in html_content

    def test_age_values_in_html(self, html_content):
        for age in self.DF["age"]:
            assert str(age) in html_content


# ---------------------------------------------------------------------------
# Test 4 — Empty DataFrame renders "No data" message
# ---------------------------------------------------------------------------

class TestEmptyDataFrame:
    """add_ae_table with 0-row DataFrame → 'No data' notice, not empty table."""

    @pytest.fixture(scope="class")
    def html_content(self, tmp_path_factory):
        out = str(tmp_path_factory.mktemp("html") / "empty.html")
        empty_df = pd.DataFrame(columns=["age", "deaths", "ae_ratio"])
        rpt = HTMLReport("Empty Test", {})
        rpt.add_ae_table(empty_df, caption="Empty table")
        rpt.render(out)
        with open(out, encoding="utf-8") as fh:
            return fh.read()

    def test_no_data_message_present(self, html_content):
        assert "No data" in html_content

    def test_no_empty_tbody(self, html_content):
        """An empty <tbody></tbody> should not appear when there is no data."""
        assert "<tbody></tbody>" not in html_content

    def test_caption_still_rendered(self, html_content):
        assert "Empty table" in html_content

    def test_excel_empty_df_writes_header_only(self, tmp_path_factory):
        """Excel: empty DataFrame writes the header row but no data rows."""
        out = str(tmp_path_factory.mktemp("xlsx") / "empty.xlsx")
        rpt = ExcelReport()
        empty_df = pd.DataFrame(columns=["age", "deaths", "ae_ratio"])
        rpt.add_ae_sheet(empty_df, sheet_name="Empty")
        rpt.render(out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Empty"]
        # Row 1 is header; row 2 should be empty
        assert ws["A1"].value == "age"
        assert ws.max_row == 1


# ---------------------------------------------------------------------------
# Test 5 — Output path / directory handling
# ---------------------------------------------------------------------------

class TestOutputPath:
    """render() creates parent directories or raises FileNotFoundError."""

    def test_creates_nested_directories(self, tmp_path):
        nested = str(tmp_path / "a" / "b" / "c" / "report.html")
        rpt = HTMLReport("Dir test", {})
        rpt.add_summary_stats({"key": "value"})
        rpt.render(nested)
        assert os.path.isfile(nested)

    def test_excel_creates_nested_directories(self, tmp_path):
        nested = str(tmp_path / "x" / "y" / "report.xlsx")
        rpt = ExcelReport()
        rpt.add_summary_sheet({"k": 1.0})
        rpt.render(nested)
        assert os.path.isfile(nested)

    def test_html_overwrites_existing_file(self, tmp_path):
        out = str(tmp_path / "report.html")
        for title in ("First", "Second"):
            rpt = HTMLReport(title, {})
            rpt.render(out)
        with open(out, encoding="utf-8") as fh:
            content = fh.read()
        assert "Second" in content
        assert "First" not in content

    def test_graduation_table_renders(self, tmp_path):
        out = str(tmp_path / "grad.html")
        rpt = HTMLReport("Grad", {})
        rpt.add_graduation_table(_grad_df(), _grad_df())
        rpt.render(out)
        with open(out, encoding="utf-8") as fh:
            content = fh.read()
        assert "Graduation Results" in content
        assert "Crude Rates" in content
        assert "Graduated Rates" in content


# ---------------------------------------------------------------------------
# Checkpoint S18 — Generate inspection artefacts
# ---------------------------------------------------------------------------

class TestCheckpointS18Artefacts:
    """Generate the HTML and Excel files for manual visual inspection.

    Files are written to the project root ``reports/`` directory so they
    persist after the test run and can be opened by the developer.
    """

    OUTPUT_DIR = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "reports",
    )

    def test_generate_html_for_inspection(self):
        """Write reports/checkpoint_s18.html for manual review."""
        df = _ae_df(10)
        rpt = HTMLReport(
            "LifeXP — Checkpoint S18 Inspection Report",
            {
                "Period":      "2020–2022",
                "Table used":  "IA 2012",
                "Methodology": "Central Ledger",
                "Generated":   "auto (pytest)",
            },
        )
        rpt.add_ae_table(df, caption="A/E by Age")
        rpt.add_graduation_table(_grad_df(10), _grad_df(10))
        rpt.add_summary_stats({
            "Overall A/E":  "1.05",
            "Total deaths": "183",
            "Total ETR":    "12 450",
            "Credibility Z": "0.87",
        })
        out = os.path.join(self.OUTPUT_DIR, "checkpoint_s18.html")
        rpt.render(out)
        assert os.path.isfile(out)

    def test_generate_excel_for_inspection(self):
        """Write reports/checkpoint_s18.xlsx for manual review."""
        df = _ae_df(10)
        rpt = ExcelReport()
        rpt.add_ae_sheet(df, sheet_name="AE by Age")
        rpt.add_ae_sheet(_ae_df(5), sheet_name="AE by Gender")
        rpt.add_summary_sheet({
            "Overall A/E":  1.05,
            "Total deaths": 183,
            "Credibility Z": 0.87,
        }, sheet_name="Summary")
        out = os.path.join(self.OUTPUT_DIR, "checkpoint_s18.xlsx")
        rpt.render(out)
        assert os.path.isfile(out)
