"""Excel experience study report using openpyxl."""

from __future__ import annotations

import os
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    numbers,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------

_RED_FILL   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri")
_DATA_FONT   = Font(name="Calibri", size=10)
_RATE_FORMAT = "0.0000"
_INT_FORMAT  = "#,##0"


# ---------------------------------------------------------------------------
# ExcelReport
# ---------------------------------------------------------------------------

class ExcelReport:
    """Openpyxl-based Excel experience study workbook.

    Sheets are accumulated via :meth:`add_ae_sheet` /
    :meth:`add_summary_sheet` and written by :meth:`render`.

    Examples
    --------
    >>> rpt = ExcelReport()
    >>> rpt.add_ae_sheet(ae_df, sheet_name="AE by Age")
    >>> rpt.add_summary_sheet({"Overall A/E": 1.05, "Deaths": 234})
    >>> rpt.render("reports/study.xlsx")
    """

    def __init__(self) -> None:
        self._wb = Workbook()
        # Remove the default empty sheet that openpyxl creates
        default = self._wb.active
        if default is not None:
            self._wb.remove(default)
        self._sheets_added: List[str] = []

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def add_ae_sheet(self, df: pd.DataFrame, sheet_name: str = "A/E") -> None:
        """Add an Actual vs Expected data sheet.

        Formatting applied:
        * Header row: bold white text on navy background.
        * Rate columns (names containing ``'rate'``, ``'ae'``, ``'qx'``,
          ``'etr'``): 4-decimal-place number format.
        * Conditional formatting on columns whose name contains ``'ae'``:
          - value > 1.2 → red fill
          - value < 0.8 → green fill
        * Column widths auto-fitted to content.

        Parameters
        ----------
        df :
            DataFrame to write.  May be empty (writes header row only).
        sheet_name :
            Worksheet name.
        """
        ws = self._wb.create_sheet(title=sheet_name)
        self._sheets_added.append(sheet_name)

        if df.empty:
            self._write_header(ws, list(df.columns))
            return

        self._write_header(ws, list(df.columns))

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, (col, val) in enumerate(zip(df.columns, row), start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = _DATA_FONT

                col_lower = str(col).lower()
                if any(k in col_lower for k in ("rate", "ae", "qx", "etr", "ratio")):
                    cell.number_format = _RATE_FORMAT
                elif isinstance(val, (int, float)) and "count" in col_lower:
                    cell.number_format = _INT_FORMAT

        # Conditional formatting on ae_ratio / ae columns
        n_rows = len(df) + 1  # +1 for header
        for c_idx, col in enumerate(df.columns, start=1):
            if "ae" in str(col).lower():
                col_letter = get_column_letter(c_idx)
                cell_range = f"{col_letter}2:{col_letter}{n_rows}"
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["1.2"],
                        fill=_RED_FILL,
                    ),
                )
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="lessThan",
                        formula=["0.8"],
                        fill=_GREEN_FILL,
                    ),
                )

        self._autofit_columns(ws, df)

    def add_summary_sheet(
        self,
        results_dict: Dict[str, Any],
        sheet_name: str = "Summary",
    ) -> None:
        """Add a two-column summary sheet (metric | value).

        Parameters
        ----------
        results_dict :
            Ordered dict of metric name → value.
        sheet_name :
            Worksheet name.
        """
        ws = self._wb.create_sheet(title=sheet_name)
        self._sheets_added.append(sheet_name)

        # Header
        for c_idx, label in enumerate(("Metric", "Value"), start=1):
            cell = ws.cell(row=1, column=c_idx, value=label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="left")

        for r_idx, (key, val) in enumerate(results_dict.items(), start=2):
            key_cell = ws.cell(row=r_idx, column=1, value=str(key))
            key_cell.font = _DATA_FONT

            val_cell = ws.cell(row=r_idx, column=2, value=val)
            val_cell.font = _DATA_FONT
            if isinstance(val, float):
                val_cell.number_format = _RATE_FORMAT

        # Auto-width
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18

    # ------------------------------------------------------------------
    # Rendering
    # ------------------------------------------------------------------

    def render(self, output_path: str) -> None:
        """Save the workbook to *output_path*.

        Creates parent directories if they do not exist.

        Parameters
        ----------
        output_path :
            Destination ``.xlsx`` path.
        """
        parent = os.path.dirname(os.path.abspath(output_path))
        os.makedirs(parent, exist_ok=True)
        self._wb.save(output_path)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_header(ws, columns: List[str]) -> None:
        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col))
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="left")

    @staticmethod
    def _autofit_columns(ws, df: pd.DataFrame) -> None:
        """Set column widths based on max content length."""
        for c_idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(c_idx)
            max_len = max(
                len(str(col)),
                df.iloc[:, c_idx - 1].astype(str).str.len().max() if not df.empty else 0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
